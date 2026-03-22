# -*- coding: utf-8 -*-
"""Excel import service."""
import openpyxl
from io import BytesIO
from sqlalchemy.orm import Session
from app.models.student import Student
from app.models.group import StudentGroup, GroupMember
from app.models.import_job import ImportJob
from app.schemas.import_job import ImportPreviewItem, ImportPreviewResponse


def generate_template() -> bytes:
    """Generate Excel template for student import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生导入模板"

    headers = ["姓名", "学号", "性别", "座号", "小组名称", "备注"]
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)

    # Add sample data
    ws.cell(2, 1, "张三")
    ws.cell(2, 2, "001")
    ws.cell(2, 3, "男")
    ws.cell(2, 4, "1")
    ws.cell(2, 5, "第一小组")
    ws.cell(2, 6, "")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _normalize_cell_text(value: object) -> str | None:
    """Convert Excel cell values to trimmed text without crashing on numbers."""
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
    elif isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    else:
        normalized = str(value).strip()
    return normalized or None


def _get_first_value(data: dict, *keys: str) -> object | None:
    """Read the first available key to support common header aliases."""
    for key in keys:
        if key in data:
            return data.get(key)
    return None


def parse_excel(file_content: bytes) -> list[dict]:
    """Parse Excel file and return list of student data"""
    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb.active

    students = []
    headers = [cell.value for cell in ws[1]]

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not any(cell.value for cell in row):
            continue

        row_data = {headers[i]: (row[i].value if i < len(headers) else None) for i in range(len(headers))}
        row_data["_row"] = row_num
        students.append(row_data)

    return students


def validate_student_data(data: dict, class_id: int, db: Session, seen_student_nos: set | None = None) -> ImportPreviewItem:
    """Validate single student data"""
    row = data.get("_row", 0)
    name = _normalize_cell_text(_get_first_value(data, "姓名")) or ""
    student_no = _normalize_cell_text(_get_first_value(data, "学号")) or ""
    gender = _normalize_cell_text(_get_first_value(data, "性别"))
    seat_no = _normalize_cell_text(_get_first_value(data, "座号", "座位号"))
    group_name = _normalize_cell_text(_get_first_value(data, "小组名称", "小组"))
    notes = _normalize_cell_text(_get_first_value(data, "备注"))

    errors = []

    if not name:
        errors.append("姓名为空")
    if not student_no:
        errors.append("学号为空")

    # Check duplicate within same batch
    if student_no and seen_student_nos is not None:
        if student_no in seen_student_nos:
            errors.append(f"学号 {student_no} 在本批次中重复")
        else:
            seen_student_nos.add(student_no)

    # Check duplicate in database
    if student_no:
        existing = db.query(Student).filter(
            Student.class_id == class_id,
            Student.student_no == student_no
        ).first()
        if existing:
            errors.append(f"学号 {student_no} 已存在")

    # Check group if provided
    if group_name:
        db.query(StudentGroup).filter(
            StudentGroup.class_id == class_id,
            StudentGroup.name == group_name
        ).first()
        # If group doesn't exist, we'll create it during import

    return ImportPreviewItem(
        row=row,
        name=name,
        student_no=student_no,
        gender=gender,
        seat_no=seat_no,
        group_name=group_name,
        notes=notes,
        status="error" if errors else "ok",
        error="; ".join(errors) if errors else None
    )


def preview_import(file_content: bytes, class_id: int, db: Session) -> ImportPreviewResponse:
    """Preview import without committing"""
    rows = parse_excel(file_content)

    items = []
    valid_count = 0
    error_count = 0
    seen_student_nos: set = set()

    for data in rows:
        item = validate_student_data(data, class_id, db, seen_student_nos)
        items.append(item)
        if item.status == "ok":
            valid_count += 1
        else:
            error_count += 1

    return ImportPreviewResponse(
        total=len(items),
        valid_count=valid_count,
        error_count=error_count,
        items=items
    )


def execute_import(file_content: bytes, class_id: int, operator_id: int, db: Session) -> ImportJob:
    """Execute actual import"""
    rows = parse_excel(file_content)

    job = ImportJob(
        class_id=class_id,
        operator_id=operator_id,
        file_name="import.xlsx",
        total_rows=len(rows),
        success_rows=0,
        failed_rows=0
    )
    db.add(job)
    db.flush()

    success_count = 0
    error_rows = []
    seen_student_nos: set = set()

    for data in rows:
        item = validate_student_data(data, class_id, db, seen_student_nos)
        if item.status != "ok":
            error_rows.append(f"第{item.row}行: {item.error}")
            continue

        # Check if group exists, create if not
        group_id = None
        if item.group_name:
            group = db.query(StudentGroup).filter(
                StudentGroup.class_id == class_id,
                StudentGroup.name == item.group_name
            ).first()
            if group:
                group_id = group.id
            else:
                new_group = StudentGroup(class_id=class_id, name=item.group_name)
                db.add(new_group)
                db.flush()
                group_id = new_group.id

        # Create student
        student = Student(
            class_id=class_id,
            name=item.name,
            student_no=item.student_no,
            gender=item.gender,
            seat_no=item.seat_no,
            group_id=group_id,
            notes=item.notes
        )
        db.add(student)
        db.flush()

        if group_id:
            db.add(GroupMember(group_id=group_id, student_id=student.id))

        success_count += 1
        seen_student_nos.add(item.student_no)

    job.success_rows = success_count
    job.failed_rows = len(error_rows)
    if error_rows:
        job.error_report = "\n".join(error_rows[:50])  # Limit to first 50 errors

    db.commit()
    db.refresh(job)
    return job
